"""Helpers sintéticos de Excel para tests (sin datos reales)."""

from __future__ import annotations

from pathlib import Path
from typing import Any


def write_xlsx(
    path: Path,
    rows: list[list[Any]],
    *,
    sheet_name: str = "Hoja1",
    extra_sheets: dict[str, list[list[Any]]] | None = None,
) -> Path:
    """Crea un .xlsx sintético."""
    from openpyxl import Workbook

    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = sheet_name
    for row in rows:
        first.append(row)
    if extra_sheets:
        for name, sheet_rows in extra_sheets.items():
            sheet = workbook.create_sheet(title=name)
            for row in sheet_rows:
                sheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


def write_xls(
    path: Path,
    rows: list[list[Any]],
    *,
    sheet_name: str = "Hoja1",
) -> Path:
    """
    Crea un .xls sintético mínimo con xlwt si está disponible;
    si no, usa una estrategia con xlrd-compatible via openpyxl export no aplica.

    Preferimos xlwt solo en tests; si no está, se omite el test de escritura
    y se generan libros con una tabla BIFF simple no — mejor instalar xlwt
    como dependencia de test o generar con estructura mínima.

    Aquí usamos ``xlwt`` opcional; el caller debe saltar si falla ImportError.
    """
    import xlwt

    book = xlwt.Workbook()
    sheet = book.add_sheet(sheet_name)
    for r_idx, row in enumerate(rows):
        for c_idx, value in enumerate(row):
            sheet.write(r_idx, c_idx, value)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(str(path))
    return path
