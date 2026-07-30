"""Lectura local de libros Excel (.xlsx / .xls) sin Microsoft Excel."""

from __future__ import annotations

import logging
import re
from collections.abc import Iterator, Sequence
from pathlib import Path
from typing import Any

from separador_nominas.constants import (
    DEPARTMENT_HEADER_ALIASES,
    DOCUMENT_HEADER_ALIASES,
    LOGGER_NAME,
    MAX_SPREADSHEET_ROWS,
    MAX_SPREADSHEET_SHEETS,
    SPREADSHEET_EXTENSIONS,
)
from separador_nominas.department_normalization import (
    normalize_department,
    to_department_key,
)
from separador_nominas.document_identifier_service import validate_document_id
from separador_nominas.exceptions import (
    InvalidSpreadsheetExtensionError,
    SpreadsheetColumnError,
    SpreadsheetEmptyError,
    SpreadsheetNotFoundError,
    SpreadsheetNotSelectedError,
    SpreadsheetProtectedError,
    SpreadsheetReadError,
    SpreadsheetSheetNotFoundError,
    SpreadsheetTooLargeError,
)
from separador_nominas.spreadsheet_models import (
    AssignmentConflict,
    ColumnMapping,
    DepartmentAssignment,
    SpreadsheetImportResult,
    SpreadsheetRowIssue,
)

logger = logging.getLogger(LOGGER_NAME)

_HEADER_NORMALIZE_RE = re.compile(r"[\s/\-_]+")


def _header_key(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    if not text:
        return None
    key = to_department_key(text)
    if key is None:
        return None
    return _HEADER_NORMALIZE_RE.sub(" ", key).strip()


def _alias_set(aliases: frozenset[str]) -> frozenset[str]:
    normalized: set[str] = set()
    for alias in aliases:
        key = _header_key(alias)
        if key:
            normalized.add(key)
    return frozenset(normalized)


_DOC_ALIASES = _alias_set(DOCUMENT_HEADER_ALIASES)
_DEPT_ALIASES = _alias_set(DEPARTMENT_HEADER_ALIASES)


def validate_spreadsheet_path(path: Path | str | None) -> Path:
    """Valida existencia y extensión del Excel."""
    if path is None or (isinstance(path, str) and not path.strip()):
        raise SpreadsheetNotSelectedError(
            "No se ha seleccionado ningún archivo Excel."
        )
    file_path = Path(path)
    if not file_path.exists():
        raise SpreadsheetNotFoundError(
            "No se ha encontrado el archivo Excel seleccionado.\n"
            "Comprueba la ruta e inténtalo de nuevo."
        )
    if not file_path.is_file():
        raise SpreadsheetNotFoundError(
            "La ruta indicada no corresponde a un archivo Excel."
        )
    suffix = file_path.suffix.lower()
    if suffix not in SPREADSHEET_EXTENSIONS:
        raise InvalidSpreadsheetExtensionError(
            "El archivo debe tener extensión .xlsx o .xls."
        )
    return file_path.resolve()


def list_sheet_names(path: Path | str) -> tuple[str, ...]:
    """Devuelve los nombres de hoja del libro (cierra el archivo al terminar)."""
    file_path = validate_spreadsheet_path(path)
    suffix = file_path.suffix.lower()
    try:
        if suffix == ".xlsx":
            return _list_sheets_xlsx(file_path)
        return _list_sheets_xls(file_path)
    except (
        SpreadsheetReadError,
        SpreadsheetProtectedError,
        SpreadsheetTooLargeError,
        SpreadsheetEmptyError,
    ):
        raise
    except Exception as exc:  # noqa: BLE001
        raise SpreadsheetReadError(
            "No se ha podido abrir el archivo Excel.\n"
            "Comprueba que no esté dañado ni abierto en exclusiva por otra aplicación."
        ) from exc


def detect_column_mapping(
    headers: Sequence[Any],
) -> ColumnMapping:
    """
    Detecta columnas por encabezados; si no hay, asume A=documento y B=departamento.
    """
    doc_idx: int | None = None
    dept_idx: int | None = None
    doc_header: str | None = None
    dept_header: str | None = None

    for index, cell in enumerate(headers):
        key = _header_key(cell)
        if key is None:
            continue
        label = " ".join(str(cell).strip().split())
        if doc_idx is None and key in _DOC_ALIASES:
            doc_idx = index
            doc_header = label
        elif dept_idx is None and key in _DEPT_ALIASES:
            dept_idx = index
            dept_header = label

    if doc_idx is not None and dept_idx is not None and doc_idx != dept_idx:
        return ColumnMapping(
            document_column_index=doc_idx,
            department_column_index=dept_idx,
            document_header=doc_header,
            department_header=dept_header,
            header_row_used=True,
        )

    if len(headers) < 2:
        raise SpreadsheetColumnError(
            "No se han encontrado dos columnas utilizables.\n"
            "Se necesitan una columna de DNI/NIE y otra de departamento."
        )

    return ColumnMapping(
        document_column_index=0,
        department_column_index=1,
        document_header=_cell_to_header_label(headers[0], "A"),
        department_header=_cell_to_header_label(headers[1], "B"),
        header_row_used=False,
    )


def _cell_to_header_label(cell: Any, fallback: str) -> str:
    if cell is None:
        return fallback
    text = " ".join(str(cell).strip().split())
    return text or fallback


def cell_value_to_document_text(value: Any) -> tuple[str | None, str | None]:
    """
    Convierte una celda a texto de documento.

    Returns:
        (texto, issue_code) — issue_code si la conversión es ambigua.
    """
    if value is None:
        return None, None
    if isinstance(value, bool):
        return None, "unsupported_cell_type"
    if isinstance(value, float):
        if value != value:  # NaN
            return None, None
        # Rechazar floats no enteros (notación científica / decimales).
        if not value.is_integer():
            return None, "scientific_or_float_ambiguous"
        # Entero almacenado como float: advertencia por posible pérdida de ceros.
        return str(int(value)), "numeric_cell_ambiguous"
    if isinstance(value, int):
        return str(value), "numeric_cell_ambiguous"
    text = str(value).strip()
    if not text:
        return None, None
    upper = text.upper().replace(" ", "")
    if re.fullmatch(r"\d+(\.\d+)?E[+\-]?\d+", upper):
        return None, "scientific_or_float_ambiguous"
    return text, None


def import_department_assignments(
    path: Path | str,
    *,
    sheet_name: str | None = None,
    column_mapping: ColumnMapping | None = None,
) -> SpreadsheetImportResult:
    """Lee asignaciones DNI/NIE → departamento desde una hoja."""
    file_path = validate_spreadsheet_path(path)
    rows, used_sheet = _read_sheet_rows(file_path, sheet_name=sheet_name)
    if not rows:
        raise SpreadsheetEmptyError(
            "La hoja seleccionada no contiene datos.\n"
            "Comprueba que el archivo tenga filas con DNI/NIE y departamento."
        )

    mapping = column_mapping or detect_column_mapping(rows[0])
    max_needed = max(mapping.document_column_index, mapping.department_column_index)
    if max_needed >= max(len(r) for r in rows):
        raise SpreadsheetColumnError(
            "Las columnas seleccionadas no existen en la hoja."
        )

    start_index = 1 if mapping.header_row_used else 0
    data_rows = rows[start_index:]
    if not data_rows:
        raise SpreadsheetEmptyError(
            "La hoja solo contiene encabezados.\n"
            "Añade filas con DNI/NIE y departamento."
        )
    if len(data_rows) > MAX_SPREADSHEET_ROWS:
        raise SpreadsheetTooLargeError(
            f"El archivo supera el límite de {MAX_SPREADSHEET_ROWS} filas.\n"
            "Reduce el contenido del Excel e inténtalo de nuevo."
        )

    warnings: list[SpreadsheetRowIssue] = []
    errors: list[SpreadsheetRowIssue] = []
    # document_id -> list of (row, dept_name, dept_key)
    by_doc: dict[str, list[tuple[int, str, str]]] = {}

    for offset, row in enumerate(data_rows):
        row_number = start_index + offset + 1  # 1-based Excel
        doc_idx = mapping.document_column_index
        dept_idx = mapping.department_column_index
        doc_raw = row[doc_idx] if doc_idx < len(row) else None
        dept_raw = row[dept_idx] if dept_idx < len(row) else None

        if _row_is_blank(doc_raw, dept_raw):
            continue

        doc_text, doc_issue = cell_value_to_document_text(doc_raw)
        if doc_issue == "scientific_or_float_ambiguous":
            errors.append(
                SpreadsheetRowIssue(
                    row_number=row_number,
                    issue_code=doc_issue,
                    user_message=(
                        f"Fila {row_number}: el documento parece estar en "
                        "notación científica o con decimales. "
                        "Formatea la columna como texto."
                    ),
                )
            )
            continue
        if doc_text is None:
            errors.append(
                SpreadsheetRowIssue(
                    row_number=row_number,
                    issue_code="empty_doc",
                    user_message=f"Fila {row_number}: falta el DNI/NIE.",
                )
            )
            continue

        match = validate_document_id(doc_text)
        if match is None or not match.format_valid:
            errors.append(
                SpreadsheetRowIssue(
                    row_number=row_number,
                    issue_code="invalid_doc",
                    user_message=(
                        f"Fila {row_number}: el DNI/NIE no tiene un formato válido."
                    ),
                )
            )
            continue
        if match.check_letter_valid is False:
            errors.append(
                SpreadsheetRowIssue(
                    row_number=row_number,
                    issue_code="invalid_check_letter",
                    user_message=(
                        f"Fila {row_number}: la letra de control del DNI/NIE "
                        "no es correcta."
                    ),
                )
            )
            continue

        if doc_issue == "numeric_cell_ambiguous":
            warnings.append(
                SpreadsheetRowIssue(
                    row_number=row_number,
                    issue_code=doc_issue,
                    user_message=(
                        f"Fila {row_number}: el documento estaba como número. "
                        "Si faltan ceros iniciales, formatea la columna como texto."
                    ),
                )
            )

        dept = normalize_department(dept_raw)
        if dept is None:
            errors.append(
                SpreadsheetRowIssue(
                    row_number=row_number,
                    issue_code="empty_dept",
                    user_message=(
                        f"Fila {row_number}: falta el departamento o el nombre "
                        "no es válido."
                    ),
                )
            )
            continue

        by_doc.setdefault(match.normalized, []).append(
            (row_number, dept.display_name, dept.department_key)
        )

    assignments: list[DepartmentAssignment] = []
    conflicts: list[AssignmentConflict] = []
    departments_order: list[str] = []
    seen_dept_keys: set[str] = set()

    for document_id, entries in by_doc.items():
        keys = {e[2] for e in entries}
        if len(keys) > 1:
            options: list[tuple[str, str]] = []
            seen_keys: set[str] = set()
            for _row, name, key in entries:
                if key not in seen_keys:
                    seen_keys.add(key)
                    options.append((key, name))
            conflicts.append(
                AssignmentConflict(
                    document_id=document_id,
                    source_rows=tuple(e[0] for e in entries),
                    department_keys=tuple(o[0] for o in options),
                    department_names=tuple(o[1] for o in options),
                    department_options=tuple(options),
                )
            )
            warnings.append(
                SpreadsheetRowIssue(
                    row_number=entries[0][0],
                    issue_code="conflicting_departments",
                    user_message=(
                        f"Filas {', '.join(str(e[0]) for e in entries)}: "
                        "el mismo DNI/NIE aparece con departamentos distintos. "
                        "Revisa el Excel o elige un departamento en la vista previa."
                    ),
                )
            )
            continue

        # Deduplicate same dept
        first = entries[0]
        if len(entries) > 1:
            warnings.append(
                SpreadsheetRowIssue(
                    row_number=first[0],
                    issue_code="duplicate_same_department",
                    user_message=(
                        f"Filas {', '.join(str(e[0]) for e in entries)}: "
                        "documento duplicado con el mismo departamento; "
                        "se mantiene una sola asignación."
                    ),
                )
            )
        assignments.append(
            DepartmentAssignment(
                document_id=document_id,
                department_name=first[1],
                department_key=first[2],
                source_row=first[0],
            )
        )
        if first[2] not in seen_dept_keys:
            seen_dept_keys.add(first[2])
            departments_order.append(first[1])

    if not assignments and not conflicts:
        if errors:
            logger.info(
                "Archivo Excel analizado sin asignaciones válidas. "
                "%s filas con errores.",
                len(errors),
            )
            return SpreadsheetImportResult(
                assignments=(),
                departments=(),
                warnings=tuple(warnings),
                errors=tuple(errors),
                conflicts=(),
                row_count_read=len(data_rows),
                sheet_name=used_sheet,
                document_column_index=mapping.document_column_index,
                department_column_index=mapping.department_column_index,
            )
        raise SpreadsheetEmptyError(
            "No se ha encontrado ninguna asignación válida en la hoja.\n"
            "Comprueba las columnas de DNI/NIE y departamento."
        )

    logger.info(
        "Archivo Excel analizado correctamente. "
        "%s filas procesadas. %s asignaciones válidas. "
        "%s departamentos detectados. %s conflictos detectados.",
        len(data_rows),
        len(assignments),
        len(departments_order),
        len(conflicts),
    )

    return SpreadsheetImportResult(
        assignments=tuple(assignments),
        departments=tuple(departments_order),
        warnings=tuple(warnings),
        errors=tuple(errors),
        conflicts=tuple(conflicts),
        row_count_read=len(data_rows),
        sheet_name=used_sheet,
        document_column_index=mapping.document_column_index,
        department_column_index=mapping.department_column_index,
    )


def _row_is_blank(doc_raw: Any, dept_raw: Any) -> bool:
    def blank(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        return False

    return blank(doc_raw) and blank(dept_raw)


def _read_sheet_rows(
    file_path: Path,
    *,
    sheet_name: str | None,
) -> tuple[list[list[Any]], str]:
    suffix = file_path.suffix.lower()
    try:
        if suffix == ".xlsx":
            return _read_rows_xlsx(file_path, sheet_name=sheet_name)
        return _read_rows_xls(file_path, sheet_name=sheet_name)
    except (
        SpreadsheetSheetNotFoundError,
        SpreadsheetEmptyError,
        SpreadsheetTooLargeError,
        SpreadsheetProtectedError,
        SpreadsheetReadError,
    ):
        raise
    except Exception as exc:  # noqa: BLE001
        raise SpreadsheetReadError(
            "No se ha podido leer el archivo Excel.\n"
            "Comprueba que no esté dañado ni protegido de forma incompatible."
        ) from exc


def _list_sheets_xlsx(file_path: Path) -> tuple[str, ...]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SpreadsheetReadError(
            "Falta la dependencia para leer archivos .xlsx (openpyxl)."
        ) from exc

    try:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001
        message = str(exc).lower()
        if "password" in message or "encrypted" in message:
            raise SpreadsheetProtectedError(
                "El archivo Excel está protegido con contraseña "
                "y no se puede abrir."
            ) from exc
        raise SpreadsheetReadError(
            "No se ha podido abrir el archivo .xlsx."
        ) from exc

    try:
        names = tuple(workbook.sheetnames)
        if not names:
            raise SpreadsheetEmptyError("El libro Excel no contiene hojas.")
        if len(names) > MAX_SPREADSHEET_SHEETS:
            raise SpreadsheetTooLargeError(
                f"El libro supera el límite de {MAX_SPREADSHEET_SHEETS} hojas."
            )
        return names
    finally:
        workbook.close()


def _list_sheets_xls(file_path: Path) -> tuple[str, ...]:
    try:
        import xlrd
    except ImportError as exc:
        raise SpreadsheetReadError(
            "Falta la dependencia para leer archivos .xls (xlrd)."
        ) from exc

    try:
        book = xlrd.open_workbook(str(file_path), on_demand=True)
    except Exception as exc:  # noqa: BLE001
        raise SpreadsheetReadError(
            "No se ha podido abrir el archivo .xls.\n"
            "Comprueba que el formato sea compatible."
        ) from exc

    try:
        names = tuple(book.sheet_names())
        if not names:
            raise SpreadsheetEmptyError("El libro Excel no contiene hojas.")
        if len(names) > MAX_SPREADSHEET_SHEETS:
            raise SpreadsheetTooLargeError(
                f"El libro supera el límite de {MAX_SPREADSHEET_SHEETS} hojas."
            )
        return names
    finally:
        book.release_resources()


def _read_rows_xlsx(
    file_path: Path,
    *,
    sheet_name: str | None,
) -> tuple[list[list[Any]], str]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )
    try:
        names = workbook.sheetnames
        if not names:
            raise SpreadsheetEmptyError("El libro Excel no contiene hojas.")
        chosen = sheet_name or names[0]
        if chosen not in names:
            raise SpreadsheetSheetNotFoundError(
                "La hoja seleccionada no existe en el archivo Excel."
            )
        sheet = workbook[chosen]
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) > MAX_SPREADSHEET_ROWS + 2:
                raise SpreadsheetTooLargeError(
                    f"El archivo supera el límite de {MAX_SPREADSHEET_ROWS} filas."
                )
        return rows, chosen
    finally:
        workbook.close()


def _read_rows_xls(
    file_path: Path,
    *,
    sheet_name: str | None,
) -> tuple[list[list[Any]], str]:
    import xlrd

    book = xlrd.open_workbook(str(file_path), on_demand=True)
    try:
        names = book.sheet_names()
        if not names:
            raise SpreadsheetEmptyError("El libro Excel no contiene hojas.")
        chosen = sheet_name or names[0]
        if chosen not in names:
            raise SpreadsheetSheetNotFoundError(
                "La hoja seleccionada no existe en el archivo Excel."
            )
        sheet = book.sheet_by_name(chosen)
        if sheet.nrows > MAX_SPREADSHEET_ROWS + 2:
            raise SpreadsheetTooLargeError(
                f"El archivo supera el límite de {MAX_SPREADSHEET_ROWS} filas."
            )
        rows: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            values: list[Any] = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                values.append(_xlrd_cell_value(book, cell))
            rows.append(values)
        return rows, chosen
    finally:
        book.release_resources()


def _xlrd_cell_value(book: Any, cell: Any) -> Any:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return None
    return cell.value


def iter_column_labels(
    headers: Sequence[Any],
    width: int | None = None,
) -> Iterator[str]:
    """Etiquetas «A - Encabezado» para la UI."""
    count = width if width is not None else max(len(headers), 2)
    for index in range(count):
        letter = _column_letter(index)
        header = ""
        if index < len(headers) and headers[index] is not None:
            header = " ".join(str(headers[index]).strip().split())
        if header:
            yield f"{letter} - {header}"
        else:
            yield letter


def _column_letter(index: int) -> str:
    """Índice 0-based → letra de columna Excel (A, B, …, Z, AA, …)."""
    n = index + 1
    letters = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def peek_header_row(
    path: Path | str,
    *,
    sheet_name: str | None = None,
) -> tuple[str, list[Any]]:
    """Lee la primera fila de una hoja para la UI de columnas."""
    file_path = validate_spreadsheet_path(path)
    rows, used = _read_sheet_rows(file_path, sheet_name=sheet_name)
    if not rows:
        raise SpreadsheetEmptyError("La hoja seleccionada no contiene datos.")
    return used, list(rows[0])
